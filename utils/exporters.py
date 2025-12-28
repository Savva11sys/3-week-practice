import pandas as pd
import json
import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class DataExporter:
    """Класс для экспорта данных"""
    
    def __init__(self, db):
        self.db = db
    
    def export_requests(self, format='excel', filters=None):
        """Экспорт заявок"""
        requests = self.db.search_requests("", filters or {})
        
        if not requests:
            return None
        
        os.makedirs('data/export', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'excel':
            return self._export_to_excel(requests, f'requests_{timestamp}.xlsx')
        elif format == 'csv':
            return self._export_to_csv(requests, f'requests_{timestamp}.csv')
        elif format == 'json':
            return self._export_to_json(requests, f'requests_{timestamp}.json')
        else:
            return None
    
    def export_statistics(self):
        """Экспорт статистики"""
        stats = self.db.get_statistics()
        
        os.makedirs('data/export', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'data/export/statistics_{timestamp}.xlsx'
        
        # Создание Excel файла
        wb = openpyxl.Workbook()
        
        # Основная статистика
        ws1 = wb.active
        ws1.title = "Общая статистика"
        
        headers = ['Показатель', 'Значение']
        data = [
            ['Всего заявок', stats['total_requests']],
            ['Активные заявки', stats['active_requests']],
            ['Завершенные заявки', stats['completed_requests']],
            ['Среднее время ремонта (дней)', f"{stats['avg_repair_days']:.1f}"],
            ['Общий доход', f"{stats['total_revenue']:,.0f}₽"],
            ['Уникальные клиенты', stats['unique_clients']]
        ]
        
        self._write_sheet(ws1, headers, data, "Общая статистика")
        
        # Статусы заявок
        if stats['by_status']:
            ws2 = wb.create_sheet("Статусы заявок")
            headers = ['Статус', 'Количество', 'Процент']
            
            total = sum(stats['by_status'].values())
            data = []
            for status, count in stats['by_status'].items():
                percentage = (count / total * 100) if total > 0 else 0
                data.append([status, count, f"{percentage:.1f}%"])
            
            self._write_sheet(ws2, headers, data, "Статусы заявок")
        
        # Типы техники
        if stats['by_tech_type']:
            ws3 = wb.create_sheet("Типы техники")
            headers = ['Тип техники', 'Количество', 'Ср. время ремонта']
            
            data = []
            for tech_type, count, avg_days in stats['by_tech_type']:
                data.append([tech_type, count, f"{avg_days or 0:.1f} дней"])
            
            self._write_sheet(ws3, headers, data, "Типы техники")
        
        # Мастера
        if stats['by_master']:
            ws4 = wb.create_sheet("Мастера")
            headers = ['Мастер', 'Всего заявок', 'Завершено', 'Ср. время']
            
            data = []
            for master_name, total, completed, avg_days in stats['by_master']:
                data.append([master_name, total, completed, f"{avg_days or 0:.1f} дней"])
            
            self._write_sheet(ws4, headers, data, "Мастера")
        
        # По месяцам
        if stats['by_month']:
            ws5 = wb.create_sheet("По месяцам")
            headers = ['Месяц', 'Всего заявок', 'Завершено', 'Процент']
            
            data = []
            for month, total, completed in stats['by_month']:
                percentage = (completed / total * 100) if total > 0 else 0
                data.append([month, total, completed, f"{percentage:.1f}%"])
            
            self._write_sheet(ws5, headers, data, "По месяцам")
        
        wb.save(filename)
        return filename
    
    def _export_to_excel(self, data, filename):
        """Экспорт в Excel"""
        df = pd.DataFrame(data)
        
        # Удаление ненужных столбцов
        columns_to_keep = [
            'requestID', 'startDate', 'homeTechType', 'homeTechModel',
            'problemDescription', 'requestStatus', 'priority',
            'estimatedCost', 'actualCost', 'completionDate'
        ]
        
        df = df[[col for col in columns_to_keep if col in df.columns]]
        
        # Переименование столбцов
        column_names = {
            'requestID': 'ID заявки',
            'startDate': 'Дата создания',
            'homeTechType': 'Тип техники',
            'homeTechModel': 'Модель',
            'problemDescription': 'Описание проблемы',
            'requestStatus': 'Статус',
            'priority': 'Приоритет',
            'estimatedCost': 'Ориентировочная стоимость',
            'actualCost': 'Фактическая стоимость',
            'completionDate': 'Дата завершения'
        }
        
        df.rename(columns=column_names, inplace=True)
        
        # Форматирование дат
        date_columns = ['Дата создания', 'Дата завершения']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.strftime('%d.%m.%Y')
        
        # Форматирование стоимости
        cost_columns = ['Ориентировочная стоимость', 'Фактическая стоимость']
        for col in cost_columns:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f"{x:,.0f}₽" if pd.notnull(x) else "")
        
        # Сохранение
        filepath = f'data/export/{filename}'
        df.to_excel(filepath, index=False, sheet_name='Заявки')
        
        # Форматирование Excel
        self._format_excel_file(filepath)
        
        return filepath
    
    def _format_excel_file(self, filepath):
        """Форматирование Excel файла"""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Стили
        header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Форматирование ячеек
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
        
        wb.save(filepath)
    
    def _export_to_csv(self, data, filename):
        """Экспорт в CSV"""
        filepath = f'data/export/{filename}'
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            if data:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
        
        return filepath
    
    def _export_to_json(self, data, filename):
        """Экспорт в JSON"""
        filepath = f'data/export/{filename}'
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        
        return filepath
    
    def _write_sheet(self, worksheet, headers, data, title):
        """Запись данных на лист Excel"""
        # Заголовок
        worksheet.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        
        # Заголовки столбцов
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Данные
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, cell_data in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_data)
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_users(self):
        """Экспорт пользователей"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM users ORDER BY type, fio')
            users = [dict(row) for row in cursor.fetchall()]
        
        if not users:
            return None
        
        os.makedirs('data/export', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'data/export/users_{timestamp}.xlsx'
        
        df = pd.DataFrame(users)
        
        # Удаление паролей
        if 'password' in df.columns:
            df = df.drop('password', axis=1)
        
        df.to_excel(filename, index=False)
        return filename
    
    def export_parts(self):
        """Экспорт запчастей"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM parts ORDER BY partName')
            parts = [dict(row) for row in cursor.fetchall()]
        
        if not parts:
            return None
        
        os.makedirs('data/export', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'data/export/parts_{timestamp}.xlsx'
        
        df = pd.DataFrame(parts)
        df.to_excel(filename, index=False)
        return filename